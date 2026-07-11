"""
Extraction de texte multi-format pour les pièces jointes d'e-mail (PDF, Word, Excel) — un vrai
appel d'offres arrive souvent avec plusieurs documents de types différents, pas un seul PDF
(P2 §11.4 item 16 : "les vrais appels d'offres = plusieurs PDF + Word + Excel ; `pdf_reader` ne
lit que le premier PDF"). `pdf_reader.py` reste inchangé pour ses usages existants (ingestion de
la Knowledge_Base, upload manuel d'un PDF unique) ; ce module gère la liste réelle de pièces
jointes d'un e-mail entrant (Gmail ou upload manuel multi-fichiers dans l'UI).
"""
import io

from docx import Document
from openpyxl import load_workbook

from .pdf_reader import MAX_CHARS, extract_raw_text_from_pdf

SUPPORTED_EXTENSIONS = ("pdf", "docx", "xlsx")


def _extract_raw_text_from_docx(data: bytes) -> str:
    """Texte brut d'un document Word (.docx), sans troncature (paragraphes uniquement)."""
    try:
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        print(f"❌ Erreur lors de l'extraction Word : {e}")
        return ""


def _extract_raw_text_from_xlsx(data: bytes) -> str:
    """Texte brut d'un classeur Excel (.xlsx), sans troncature (toutes les feuilles, ligne par ligne)."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    lines.append(" | ".join(cells))
        return "\n".join(lines)
    except Exception as e:
        print(f"❌ Erreur lors de l'extraction Excel : {e}")
        return ""


_EXTRACTORS = {
    "pdf": extract_raw_text_from_pdf,
    "docx": _extract_raw_text_from_docx,
    "xlsx": _extract_raw_text_from_xlsx,
}


def extract_text_from_attachments(attachments: list[tuple[str, bytes]]) -> str:
    """
    Concatène le texte de plusieurs pièces jointes (PDF/Word/Excel), chacune préfixée par son nom
    de fichier pour que le LLM sache à quoi correspond chaque section, puis tronque l'ENSEMBLE à
    MAX_CHARS (un seul budget global par e-mail — pas un budget par fichier, pour ne pas exploser
    le contexte LLM avec 5 pièces jointes). Extension non supportée ou extraction échouée :
    ignorée silencieusement (dégradation gracieuse, même principe que le reste du projet).
    """
    parts = []
    for filename, data in attachments:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        extractor = _EXTRACTORS.get(ext)
        if not extractor:
            continue
        text = extractor(data)
        if text:
            parts.append(f"--- Pièce jointe : {filename} ---\n{text}")

    combined = "\n\n".join(parts)
    if len(combined) > MAX_CHARS:
        combined = combined[:MAX_CHARS] + "... [Texte tronqué car trop long]"
    return combined
